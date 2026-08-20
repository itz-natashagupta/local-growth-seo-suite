"""
Google Maps Lead Scraper — Core Engine
Scrapes business leads from Google Maps based on category and city,
filters by website availability if requested, and exports formatted Excel files.
"""
import os
import random
import re
import sys
import time
from datetime import datetime

import pandas as pd
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

MAX_RESULTS = 50
SCROLL_PAUSE = 0.2
ACTION_DELAY = (0.2, 0.5)
HEADLESS = True


def human_delay(low: float = ACTION_DELAY[0], high: float = ACTION_DELAY[1]) -> None:
    time.sleep(random.uniform(low, high))


def safe_text(locator, timeout: int = 1500) -> str:
    try:
        return locator.first.inner_text(timeout=timeout).strip()
    except Exception:
        return ""


def safe_attribute(locator, attr: str, timeout: int = 1500) -> str:
    try:
        return (locator.first.get_attribute(attr, timeout=timeout) or "").strip()
    except Exception:
        return ""


# ─── Coaching-Class Intent Engine ────────────────────────────────────────────

def _get_best_call_window(hours_text: str) -> str:
    """Suggest the optimal cold-call time window based on business opening hours."""
    h = (hours_text or "").lower()
    if re.search(r'open 24|24/7|24 hrs', h):
        return "10:00 AM – 11:00 AM (Morning sweet-spot, before peak hours)"
    if re.search(r'9:00\s*am|9\s*am|9:30\s*am', h):
        return "10:30 AM – 12:00 PM (After morning batch, before lunch)"
    if re.search(r'10:00\s*am|10\s*am|10:30\s*am', h):
        return "11:00 AM – 12:30 PM (Mid-morning sweet-spot)"
    if re.search(r'4:00\s*pm|4\s*pm|5\s*pm|evening', h):
        return "12:00 PM – 1:00 PM (Owner free before evening batches)"
    return "10:00 AM – 12:00 PM (Best general cold-call window)"


def _build_cold_call_pitch(biz_name, rtg, revs, has_web, is_http_only, has_phone):
    """Return a 30-second coaching-class specific cold call pitch script."""
    if not has_web and rtg >= 4.0 and revs >= 10:
        return (f"Hi, am I speaking with the owner of {biz_name}? Great! "
                f"I was looking up coaching classes in your area and saw your institute has {rtg} stars "
                f"with {revs} parent reviews on Google — that is really impressive. "
                f"But I noticed there is no website, which means parents searching at 11 PM "
                f"cannot find your fee structure or batch schedule and they end up enrolling at another institute. "
                f"We build coaching institute websites in 48 hours with online inquiry forms, batch info, and parent testimonials. "
                f"Can I send you a 1-minute video of what it would look like for {biz_name}?")
    if not has_web:
        return (f"Hi, is this {biz_name}? I specialize in helping coaching institutes get more student admissions through Google. "
                f"I noticed your institute does not have a website, which means parents who search online at night "
                f"cannot find your batches or fees and they go to a competitor. "
                f"We build institute websites with inquiry forms in under 48 hours. "
                f"Would you be open to a 2-minute chat?")
    if is_http_only:
        return (f"Hi, this is regarding {biz_name}'s website. When parents click your site from Google Maps, "
                f"their browser shows a big red 'Not Secure' warning. Research shows 70 percent of parents "
                f"immediately close that tab and go to a competitor instead — especially when enrolling their child. "
                f"We fix this SSL security issue in 24 hours. "
                f"Would you like me to send you a screenshot of what parents currently see?")
    if revs < 15:
        return (f"Hi, am I speaking with the owner of {biz_name}? "
                f"I help coaching institutes rank at the top of Google Maps and get more student inquiries. "
                f"I noticed your institute only has {revs} reviews, while top institutes in your area have 80 to 100. "
                f"Parents almost always pick the coaching with more reviews. "
                f"We set up an automated system that collects reviews from every student's parent after each batch. "
                f"Can I share how this works in 3 minutes?")
    if not has_phone:
        return (f"Hi, I noticed {biz_name}'s Google Maps listing does not show a contact number for parents to call. "
                f"This means you are losing direct admission inquiries every single day. "
                f"This is a quick fix that takes under 1 hour. "
                f"Would you like our team to help you update this listing today?")
    return (f"Hi, I help coaching institutes like {biz_name} rank number 1 on Google Maps "
            f"and get 20 to 30 extra admission inquiries every month. "
            f"We have worked with similar institutes in your city. "
            f"Would you be open to a free 5-minute digital audit where I show you exactly "
            f"what is stopping more parents from finding you online?")


def _build_whatsapp_message(biz_name, rtg, revs, has_web):
    """Return a WhatsApp Business intro message for coaching class owners."""
    if not has_web:
        return (f"Hello Sir/Ma'am \U0001f64f\n\n"
                f"I came across *{biz_name}* on Google Maps and was impressed by your ratings.\n\n"
                f"I noticed you don't have a website yet, which means parents searching online late at night "
                f"can't find your batch details or fee structure — and they end up enrolling at another institute.\n\n"
                f"We specialize in building *coaching institute websites* with online inquiry forms, "
                f"results showcase, and batch schedules — ready in just 48 hours \u26a1\n\n"
                f"Can I share a 1-min demo video? No charges for the call \U0001f60a")
    if revs < 20:
        return (f"Hello Sir/Ma'am \U0001f64f\n\n"
                f"This is regarding *{biz_name}* on Google Maps.\n\n"
                f"Parents compare institutes based on Google reviews before enrolling their children. "
                f"Your institute currently has only *{revs} reviews*, while top institutes in your area have 80+.\n\n"
                f"We help coaching institutes get *50+ genuine Google reviews* from existing students "
                f"using an automated follow-up system \U0001f4f2\n\n"
                f"Takes less than 30 days. Can we do a quick 5-min call to explain?")
    return (f"Hello Sir/Ma'am \U0001f64f\n\n"
            f"I came across *{biz_name}* while researching coaching institutes in your area.\n\n"
            f"I help institutes rank #1 on Google Maps and get more admission inquiries every month through Local SEO.\n\n"
            f"Would you be open to a *free 5-minute digital audit* to see how you compare to top competitors? \U0001f4ca\n\n"
            f"No obligation, just useful insights for your institute \U0001f60a")


def _build_email_template(biz_name, rtg, revs, has_web):
    """Return (subject, body) for a coaching-class follow-up cold email."""
    if not has_web:
        subject = f"More Student Admissions for {biz_name} — Website in 48 Hours"
        body = (f"Dear Director,\n\n"
                f"I recently came across {biz_name} on Google Maps and noticed your institute has excellent parent reviews.\n\n"
                f"However, there is no website listed for your institute. In today's digital age, most parents "
                f"search for coaching classes online — especially late at night when comparing options for their children. "
                f"Without a website, you are missing out on these admission inquiries entirely.\n\n"
                f"Our team specialises in building professional coaching institute websites that include:\n"
                f"  • Online admission inquiry forms\n"
                f"  • Batch schedules and fee structure pages\n"
                f"  • Student results and testimonials section\n"
                f"  • Mobile-friendly design optimised for Google Search\n\n"
                f"We deliver within 48 hours. I would love to send you a free sample website we built for a similar institute in your city.\n\n"
                f"Would a 10-minute call this week work for you?\n\n"
                f"Best regards,\n[Your Name]\n[Your Agency Name]\n[Your Phone Number]")
    elif revs < 20:
        subject = f"More Student Admissions for {biz_name} — Free Google Maps Audit"
        body = (f"Dear Director,\n\n"
                f"I noticed that {biz_name} currently has {revs} reviews on Google Maps. "
                f"While your rating is good, top coaching institutes in your area have 80 to 100+ reviews "
                f"and parents almost always choose the institute with more social proof.\n\n"
                f"We help coaching institutes collect 50+ genuine Google reviews from existing students "
                f"and parents within 30 days using a simple automated follow-up system.\n\n"
                f"This directly leads to:\n"
                f"  • Higher ranking in Google Maps Top 3 results\n"
                f"  • More parent trust and admission inquiries\n"
                f"  • Better conversion when parents compare institutes\n\n"
                f"I would love to share a free audit comparing your current presence vs your top competitor.\n\n"
                f"Best regards,\n[Your Name]\n[Your Agency Name]\n[Your Phone Number]")
    else:
        subject = f"Free Digital Audit for {biz_name} — Rank #1 in Your City"
        body = (f"Dear Director,\n\n"
                f"I am reaching out regarding {biz_name}'s online presence on Google Maps.\n\n"
                f"We help coaching institutes rank #1 on Google Maps and increase student admissions "
                f"by 30 to 50 percent through targeted Local SEO strategies.\n\n"
                f"I would like to offer a completely free 5-minute digital audit that shows:\n"
                f"  • How your institute appears vs top competitors in your city\n"
                f"  • Exact gaps in your Google Business Profile\n"
                f"  • Quick wins that can increase inquiries within 30 days\n\n"
                f"Would you be open to a quick call this week?\n\n"
                f"Best regards,\n[Your Name]\n[Your Agency Name]\n[Your Phone Number]")
    return subject, body


def analyze_lead_intent(lead: dict) -> dict:
    """Multi-trigger stacking intent engine — coaching class edition.
    Returns a rich dict with tier, badge, all pain points, best call window,
    cold call pitch, WhatsApp message, and email template.
    """
    biz_name = lead.get("Business Name", "there")
    has_web  = bool(lead.get("Website") and lead.get("Website").strip() not in ("", "N/A"))
    web_url  = lead.get("Website", "").strip()
    hours_text = lead.get("Opening Hours", "N/A")

    try:  rtg  = float(lead.get("Rating") or 0)
    except Exception: rtg = 0.0

    try:  revs = int(lead.get("Number of Reviews") or 0)
    except Exception: revs = 0

    has_phone    = bool(lead.get("Phone Number") and lead.get("Phone Number").strip() not in ("", "N/A"))
    is_http_only = has_web and web_url.startswith("http://")

    # ── Multi-trigger stacking: accumulate ALL pain points ──────────────────
    score = 50
    pain_points = []

    if not has_web:
        score += 40
        if rtg >= 4.0 and revs >= 10:
            score += 8
            pain_points.append(f"No Website despite {rtg}\u2605 reputation & {revs} reviews")
        else:
            pain_points.append("No Website — invisible to parents searching online")

    if is_http_only:
        score += 18
        pain_points.append("Unsecured HTTP site — Chrome 'Not Secure' warning scares parents")

    if revs < 10:
        score += 18
        pain_points.append(f"Very low reviews ({revs}) — parents won't trust without social proof")
    elif revs < 40:
        score += 10
        pain_points.append(f"Low reviews ({revs}) vs competitors with 80–100+")

    if not has_phone:
        score += 15
        pain_points.append("No phone number on Google Maps — losing direct admission calls")

    if 0 < rtg < 4.0:
        score += 5
        pain_points.append(f"Below-average rating ({rtg}\u2605) — needs reputation management")

    score = min(score, 98)

    # ── Tier & badge ────────────────────────────────────────────────────────
    if score >= 80:
        badge = f"\U0001f525 {score}% HOT LEAD"
        tier  = "\U0001f170\ufe0f Tier A \u2014 Close Today"
    elif score >= 65:
        badge = f"\u26a1 {score}% WARM LEAD"
        tier  = "\U0001f171\ufe0f Tier B \u2014 Follow Up in 2 Days"
    else:
        badge = f"\u2744\ufe0f {score}% COLD LEAD"
        tier  = "\U0001f172 Tier C \u2014 Nurture (Email Campaign)"

    primary_pain  = pain_points[0] if pain_points else "General Local SEO Optimisation"
    all_pains_str = " | ".join(pain_points) if pain_points else primary_pain

    best_call   = _get_best_call_window(hours_text)
    pitch       = _build_cold_call_pitch(biz_name, rtg, revs, has_web, is_http_only, has_phone)
    whatsapp    = _build_whatsapp_message(biz_name, rtg, revs, has_web)
    email_subj, email_body = _build_email_template(biz_name, rtg, revs, has_web)

    return {
        "raw_score":      score,
        "badge":          badge,
        "tier":           tier,
        "primary_pain":   primary_pain,
        "all_pains":      all_pains_str,
        "best_call_window": best_call,
        "pitch":          pitch,
        "whatsapp":       whatsapp,
        "email_subject":  email_subj,
        "email_body":     email_body,
    }


def extract_lead(page) -> dict:
    lead = {
        "Business Name":        "",
        "Rating":               "",
        "Number of Reviews":    "",
        "Address":              "",
        "Phone Number":         "",
        "Opening Hours":        "",
        "Is 24/7":              "No",
        "Website":              "",
        "Conversion Score":     "",
        "Raw Score":            50,
        "Tier":                 "",
        "Primary Pain Point":   "",
        "All Pain Points":      "",
        "Best Call Window":     "",
        "Cold Call Pitch Script": "",
        "WhatsApp Message":     "",
        "Email Subject":        "",
        "Follow-Up Email":      "",
        "Google Maps Link":     page.url,
    }

    try:
        lead["Business Name"] = safe_text(page.locator('h1.DUwDvf, h1[class*="fontHeadlineLarge"]'))

        rating, reviews = page.evaluate("""() => {
            const feed = document.querySelector('div[role="feed"]');
            function inFeed(el) { return feed ? feed.contains(el) : false; }
            const allEls = Array.from(document.querySelectorAll('[aria-label]'));
            const allSpans = Array.from(document.querySelectorAll('span'));
            let rating = '';
            let reviews = '';

            for (const el of allEls) {
                if (inFeed(el)) continue;
                const label = el.getAttribute('aria-label') || '';
                const m = label.match(/([1-5][.,][0-9])\\s*stars?/i);
                if (m) { rating = m[1].replace(',', '.'); break; }
            }
            if (!rating) {
                for (const span of allSpans) {
                    if (inFeed(span)) continue;
                    const txt = span.textContent.strip ? span.textContent.strip() : span.textContent.trim();
                    if (/^[1-5]\\.[0-9]$/.test(txt)) { rating = txt; break; }
                }
            }

            for (const el of allEls) {
                if (inFeed(el)) continue;
                const label = el.getAttribute('aria-label') || '';
                const m = label.match(/^([ \\d,]+)\\s*reviews?$/i);
                if (m) { reviews = m[1].replace(/[, ]/g, ''); break; }
            }
            if (!reviews) {
                for (const el of allEls) {
                    if (inFeed(el)) continue;
                    const label = el.getAttribute('aria-label') || '';
                    const m = label.match(/([\\d,]+)\\s*reviews?/i);
                    if (m) { reviews = m[1].replace(/,/g, ''); break; }
                }
            }
            if (!reviews) {
                for (const span of allSpans) {
                    if (inFeed(span)) continue;
                    const txt = span.textContent.trim();
                    const m = txt.match(/^\\(([\\d,]+)\\)$/);
                    if (m) { reviews = m[1].replace(/,/g, ''); break; }
                }
            }
            return [rating, reviews];
        }""")

        lead["Rating"] = rating
        lead["Number of Reviews"] = reviews

        address_el = page.locator('[data-item-id="address"] .Io6YTe, [data-tooltip="Copy address"] .Io6YTe')
        lead["Address"] = safe_text(address_el)

        phone_el = page.locator('[data-item-id^="phone:"] .Io6YTe')
        lead["Phone Number"] = safe_text(phone_el)

        # Opening Hours & 24/7 extraction
        hours_el = page.locator('[data-item-id*="oh"] .Io6YTe, [aria-label*="hours"], [data-tooltip*="hours"]')
        hours_text = safe_text(hours_el)
        if not hours_text:
            hours_text = safe_attribute(page.locator('[aria-label*="hours"]'), "aria-label")
        if not hours_text:
            try:
                hours_text = page.evaluate("""() => {
                    const els = Array.from(document.querySelectorAll('*'));
                    for (const el of els) {
                        const txt = (el.textContent || '').trim();
                        if (/open 24 hours|24 hours/i.test(txt) && txt.length < 40) {
                            return txt;
                        }
                    }
                    return '';
                }""")
            except Exception:
                hours_text = ""

        lead["Opening Hours"] = hours_text if hours_text else "N/A"

        page_text = ""
        try:
            page_text = page.locator('body').inner_text(timeout=1000).lower()
        except Exception:
            pass

        is_24_7_match = bool(
            re.search(r'open 24 hours|24 hours|24/7|open 24 hrs|24 hrs', hours_text.lower()) or
            re.search(r'open 24 hours|24 hours|24/7', page_text)
        )
        lead["Is 24/7"] = "Yes (24/7)" if is_24_7_match else "No"

        website_el = page.locator('[data-item-id="authority"] a, [aria-label^="Website:"] a')
        lead["Website"] = safe_attribute(website_el, "href")
        if not lead["Website"]:
            lead["Website"] = safe_text(page.locator('[data-item-id="authority"] .Io6YTe'))

        if lead["Website"] and not lead["Website"].startswith("http"):
            lead["Website"] = "https://" + lead["Website"]

        maps_url = page.url
        if "google.com/maps/place" in maps_url:
            lead["Google Maps Link"] = maps_url.split("?")[0]

        intent = analyze_lead_intent(lead)
        lead["Conversion Score"]     = intent["badge"]
        lead["Raw Score"]            = intent["raw_score"]
        lead["Tier"]                 = intent["tier"]
        lead["Primary Pain Point"]   = intent["primary_pain"]
        lead["All Pain Points"]      = intent["all_pains"]
        lead["Best Call Window"]     = intent["best_call_window"]
        lead["Cold Call Pitch Script"] = intent["pitch"]
        lead["WhatsApp Message"]     = intent["whatsapp"]
        lead["Email Subject"]        = intent["email_subject"]
        lead["Follow-Up Email"]      = intent["email_body"]

    except Exception as e:
        print(f"  [WARN] Error extracting details: {e}")

    return lead


def scrape_google_maps(category: str, city: str, max_results: int = MAX_RESULTS,
                       progress_callback=None, only_no_website: bool = False,
                       only_24_7: bool = False, only_hot_leads: bool = False) -> list:
    def emit(msg):
        if progress_callback:
            progress_callback(msg)
        else:
            print(msg)

    search_query = f"{category} in {city}"
    search_url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
    leads = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--start-maximized",
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--ignore-certificate-errors",
                "--host-resolver-rules=",
            ]
        )
        context = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
        )
        page = context.new_page()

        filters = []
        if only_no_website: filters.append("NO WEBSITE ONLY")
        if only_24_7: filters.append("OPEN 24/7 ONLY")
        if only_hot_leads: filters.append("🔥 HOT LEADS ONLY (≥65% Intent)")
        filter_msg = f" [FILTERS: {', '.join(filters)}]" if filters else ""

        emit(f"[SEARCH] Searching Google Maps for: '{search_query}'{filter_msg}")
        emit(f"   URL: {search_url}")

        page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
        human_delay(2, 2.5)

        try:
            reject_btn = page.locator('button:has-text("Reject all"), button:has-text("Accept all")')
            if reject_btn.count() > 0:
                reject_btn.first.click()
                emit("   [OK] Handled cookie consent dialog.")
        except Exception:
            pass

        try:
            page.wait_for_selector('div[role="feed"]', timeout=15000)
        except PlaywrightTimeoutError:
            emit("   [WARN] Results panel did not appear.")
            browser.close()
            return leads

        try:
            page.wait_for_selector('a.hfpxzc', timeout=12000)
            emit("   [OK] Listings loaded. Starting collection...")
        except PlaywrightTimeoutError:
            emit("   [WARN] No listing cards found.")
            browser.close()
            return leads

        target_url_count = (max_results * 4) if ((only_no_website or only_24_7 or only_hot_leads) and max_results < 9000) else max_results
        emit(f"[SCROLL] Collecting business links (target: {target_url_count})...")
        business_urls = []
        previous_count = 0
        no_new_count   = 0

        while len(business_urls) < target_url_count:
            cards = page.locator('a.hfpxzc').all()
            seen = set(business_urls)
            for card in cards:
                try:
                    href = card.get_attribute("href") or ""
                    if href and href not in seen:
                        business_urls.append(href)
                        seen.add(href)
                except Exception:
                    continue

            emit(f"   Found {len(business_urls)} links so far...")

            if page.locator('span.HlvSq').count() and page.locator('span.HlvSq').is_visible():
                emit(f"   [DONE] End of results. Total links: {len(business_urls)}")
                break

            if len(business_urls) == previous_count:
                no_new_count += 1
                if no_new_count >= 8:
                    emit(f"   [DONE] No new links after scrolling. Total: {len(business_urls)}")
                    break
                time.sleep(1.5)
            else:
                no_new_count = 0

            previous_count = len(business_urls)

            try:
                page.locator('div[role="feed"]').evaluate("el => el.scrollBy(0, 2500)")
            except Exception:
                page.keyboard.press("End")
            time.sleep(SCROLL_PAUSE)

        business_urls = business_urls[:target_url_count]
        emit(f"[OK] Collected {len(business_urls)} business URLs. Visiting each details page...")

        for idx, url in enumerate(business_urls, start=1):
            if len(leads) >= max_results:
                emit(f"[DONE] Collected target of {max_results} leads!")
                break
            try:
                emit(f"[{idx:02d}/{len(business_urls)}] Fetching lead details...")
                page.goto(url, wait_until="domcontentloaded", timeout=30000)

                try:
                    page.wait_for_selector("h1", timeout=8000)
                except PlaywrightTimeoutError:
                    pass

                lead = extract_lead(page)
                has_website = bool(lead["Website"] and lead["Website"].strip() not in ("", "N/A"))

                if only_no_website and has_website:
                    emit(f"   [SKIP] Has website ({lead['Website']}): {lead['Business Name']}")
                    human_delay(*ACTION_DELAY)
                    continue

                if only_24_7 and lead["Is 24/7"] != "Yes (24/7)":
                    emit(f"   [SKIP] Not open 24/7 ({lead['Opening Hours']}): {lead['Business Name']}")
                    human_delay(*ACTION_DELAY)
                    continue

                if only_hot_leads and lead.get("Raw Score", 50) < 65:
                    emit(f"   [SKIP] Low Conversion Intent ({lead['Conversion Score']}): {lead['Business Name']}")
                    human_delay(*ACTION_DELAY)
                    continue

                leads.append(lead)
                emit(f"   DONE: {lead['Business Name']} | {lead['Conversion Score']} | {lead['Rating']} ★ | 📞 {lead['Phone Number']}")
                human_delay(*ACTION_DELAY)

            except Exception as e:
                emit(f"  [{idx:02d}] [ERROR] {e}")
                continue

        browser.close()

    return leads


def export_to_excel(leads: list, category: str, city: str) -> str:
    if not leads:
        print("\n[WARN] No leads to export.")
        return ""

    df = pd.DataFrame(leads, columns=[
        "Business Name", "Conversion Score", "Tier", "Primary Pain Point", "All Pain Points",
        "Best Call Window", "Rating", "Number of Reviews", "Address", "Phone Number",
        "Opening Hours", "Is 24/7", "Website", "Cold Call Pitch Script",
        "WhatsApp Message", "Email Subject", "Follow-Up Email", "Google Maps Link",
    ])

    for col in ["Phone Number", "Website", "Rating", "Number of Reviews", "Cold Call Pitch Script"]:
        df[col] = df[col].replace("", "N/A").fillna("N/A")

    df = df.drop_duplicates(subset=["Business Name", "Google Maps Link"], keep="first").reset_index(drop=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_cat = re.sub(r'[^\w]', '_', category)
    safe_city = re.sub(r'[^\w]', '_', city)
    filename = f"leads_{safe_cat}_{safe_city}_{timestamp}.xlsx"
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        worksheet = writer.sheets["Leads"]

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        light_fill = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center", wrap_text=False)

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = light_fill if row_idx % 2 == 0 else PatternFill()
            for cell in row:
                cell.font = data_font
                cell.fill = fill
                cell.alignment = data_align
                cell.border = thin_border

        col_widths = {
            "A": 26, "B": 18, "C": 22, "D": 32, "E": 42, "F": 30,
            "G": 10, "H": 14, "I": 30, "J": 16, "K": 22, "L": 12,
            "M": 35, "N": 55, "O": 55, "P": 32, "Q": 65, "R": 48,
        }
        for col_letter, width in col_widths.items():
            try:
                worksheet.column_dimensions[col_letter].width = width
            except Exception:
                pass

        worksheet.freeze_panes = "A2"

        # Hyperlink website column
        for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns)):
            for cell in row:
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(name="Calibri", size=10, color="1A73E8", underline="single")

        # Color-code by tier in column B (Conversion Score)
        hot_fill  = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
        warm_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        score_col_idx = list(df.columns).index("Conversion Score") + 1
        for row in worksheet.iter_rows(min_row=2):
            val = str(row[score_col_idx - 1].value or "")
            if "HOT" in val:
                for cell in row:
                    cell.fill = hot_fill
            elif "WARM" in val:
                for cell in row:
                    cell.fill = warm_fill

        worksheet.auto_filter.ref = worksheet.dimensions

        # ─── Sheet 2: Tier A Hit-List ──────────────────────────────────────────
        raw_scores = [lead.get("Raw Score", 50) for lead in leads]
        tier_a_leads = [l for l, s in zip(leads, raw_scores) if s >= 80]
        tier_a_leads.sort(key=lambda l: l.get("Raw Score", 50), reverse=True)

        if tier_a_leads:
            df_a = pd.DataFrame(tier_a_leads, columns=[
                "Business Name", "Conversion Score", "Tier", "All Pain Points",
                "Best Call Window", "Rating", "Number of Reviews", "Phone Number",
                "Website", "Address", "Google Maps Link",
            ])
            df_a.to_excel(writer, index=False, sheet_name="🅰️ Tier A Hit List")
            ws_a = writer.sheets["🅰️ Tier A Hit List"]
            hot_hdr_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
            for cell in ws_a[1]:
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                cell.fill = hot_hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            for row_idx, row in enumerate(ws_a.iter_rows(min_row=2), start=2):
                fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid") if row_idx % 2 == 0 else PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
                for cell in row:
                    cell.font = Font(name="Calibri", size=10)
                    cell.fill = fill
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border
            a_col_widths = {"A": 26, "B": 18, "C": 22, "D": 50, "E": 32, "F": 10, "G": 14, "H": 16, "I": 35, "J": 30, "K": 50}
            for col_letter, width in a_col_widths.items():
                try:
                    ws_a.column_dimensions[col_letter].width = width
                except Exception:
                    pass
            ws_a.freeze_panes = "A2"
            ws_a.auto_filter.ref = ws_a.dimensions

        # ─── Sheet 3: Sales Scripts ─────────────────────────────────────────────
        df_scripts = pd.DataFrame(leads, columns=[
            "Business Name", "Tier", "Best Call Window",
            "Cold Call Pitch Script", "WhatsApp Message", "Email Subject", "Follow-Up Email",
        ])
        df_scripts.to_excel(writer, index=False, sheet_name="📞 Sales Scripts")
        ws_s = writer.sheets["📞 Sales Scripts"]
        script_hdr_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        for cell in ws_s[1]:
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            cell.fill = script_hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for row_idx, row in enumerate(ws_s.iter_rows(min_row=2), start=2):
            fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
            for cell in row:
                cell.font = Font(name="Calibri", size=10)
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
        s_col_widths = {"A": 26, "B": 22, "C": 30, "D": 65, "E": 60, "F": 42, "G": 65}
        for col_letter, width in s_col_widths.items():
            try:
                ws_s.column_dimensions[col_letter].width = width
            except Exception:
                pass
        ws_s.row_dimensions[1].height = 25
        for i in range(2, len(df_scripts) + 2):
            ws_s.row_dimensions[i].height = 80
        ws_s.freeze_panes = "A2"

    print(f"\n[OK] Excel saved: {filepath}")
    return filepath
