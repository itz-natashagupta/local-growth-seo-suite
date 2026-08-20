"""
Local SEO Analyzer — Core Analysis Engine
Analyzes Google Business Profile + Website SEO + Local SEO factors
Generates a professional Excel report with scores and recommendations.
"""
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

HEADLESS = True
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    )
}


def scrape_gbp(business_name: str, category: str, location: str,
               progress_callback=None) -> dict:
    def emit(msg):
        if progress_callback:
            progress_callback(msg)
        else:
            print(msg)

    gbp = {
        "Business Name": business_name,
        "Category": category,
        "Rating": "N/A",
        "Total Reviews": "N/A",
        "Website": "N/A",
        "Phone": "N/A",
        "Address": "N/A",
        "Business Hours": "Not listed",
        "Google Maps Link": "N/A",
    }

    search_query = f"{business_name} {category} {location}"
    search_url   = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"

    emit(f"[GBP] Searching Google Maps: '{search_query}'")

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=HEADLESS,
                args=["--no-sandbox", "--disable-dev-shm-usage",
                      "--disable-blink-features=AutomationControlled"]
            )
            ctx  = browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=HEADERS["User-Agent"],
                locale="en-US",
            )
            page = ctx.new_page()
            page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
            time.sleep(3)

            try:
                for txt in ["Accept all", "Reject all", "I agree"]:
                    btn = page.locator(f'button:has-text("{txt}")')
                    if btn.count() > 0:
                        btn.first.click()
                        time.sleep(1)
                        break
            except Exception:
                pass

            if "/place/" in page.url:
                emit("[GBP] Found direct business page.")
                time.sleep(2)
            else:
                card_found = False
                card_selectors = [
                    "a.hfpxzc", "a[href*='/maps/place/']",
                    "div[role='feed'] a[href]", ".Nv2PK a",
                    "div[jsaction*='mouseover'] a",
                ]
                for sel in card_selectors:
                    try:
                        page.wait_for_selector(sel, timeout=6000)
                        page.locator(sel).first.click()
                        card_found = True
                        emit(f"[GBP] Found results using selector: {sel}")
                        break
                    except Exception:
                        continue

                if not card_found:
                    emit("[GBP] No results found on Google Maps. Skipping GBP data.")
                    browser.close()
                    return gbp

                time.sleep(3.5)

            try:
                page.wait_for_selector("h1", timeout=10000)
            except PlaywrightTimeoutError:
                pass

            time.sleep(2)
            gbp["Google Maps Link"] = page.url.split("?")[0]

            try:
                raw_name = page.locator("h1").first.inner_text(timeout=3000).strip()
                if raw_name and raw_name.lower() not in ("results", "google maps", "search results", "search", "all", ""):
                    gbp["Business Name"] = raw_name
                else:
                    gbp["Business Name"] = business_name
            except Exception:
                gbp["Business Name"] = business_name

            extracted = page.evaluate("""() => {
                const d = {rating:null, reviews:null, phone:null, address:null, website:null, hours:null};
                document.querySelectorAll('[aria-label]').forEach(el => {
                    const lbl = (el.getAttribute('aria-label') || '').trim();
                    if (!d.rating && /\\d\\.\\d.*star/i.test(lbl)) {
                        const m = lbl.match(/(\\d\\.\\d)/);
                        if (m) d.rating = m[1];
                    }
                    if (!d.reviews && /[\\d,]+\\s+reviews?/i.test(lbl)) {
                        const m = lbl.match(/([\\d,]+)\\s+reviews?/i);
                        if (m) d.reviews = m[1].replace(/,/g,'');
                    }
                });

                document.querySelectorAll('[data-item-id]').forEach(el => {
                    const id = (el.getAttribute('data-item-id') || '').toLowerCase();
                    const txt = el.textContent.trim();
                    if (!d.phone && id.startsWith('phone')) {
                        const m = txt.match(/[\\+\\d][\\d\\s\\-\\(\\)]{7,}/);
                        if (m) d.phone = m[0].trim();
                    }
                    if (!d.address && id === 'address') {
                        d.address = txt.split('\\n')[0].trim();
                    }
                    if (!d.website && id === 'authority') {
                        const a = el.tagName === 'A' ? el : el.querySelector('a');
                        if (a && a.href && !a.href.includes('google.com')) d.website = a.href;
                    }
                    if (!d.hours && id === 'oh') d.hours = txt.split('\\n')[0].trim();
                });

                if (!d.website) {
                    document.querySelectorAll('a[href]').forEach(el => {
                        const lbl = (el.getAttribute('aria-label') || '').toLowerCase();
                        const href = el.href || '';
                        if (!d.website && lbl.includes('website') && href && !href.includes('google.com') && href.startsWith('http')) {
                            d.website = href;
                        }
                    });
                }
                const bodyTxt = document.body.innerText;
                if (!d.rating) {
                    const m = bodyTxt.match(/(\\d\\.\\d)\\s*(?:stars?|★)/i);
                    if (m) d.rating = m[1];
                }
                if (!d.reviews) {
                    const m = bodyTxt.match(/([\\d,]+)\\s*reviews?/i);
                    if (m) d.reviews = m[1].replace(/,/g,'');
                }
                if (!d.phone) {
                    const m = bodyTxt.match(/(\\+91[\\s\\-]?\\d{10}|\\b0\\d{2,4}[\\s\\-]\\d{6,8}|\\b\\d{10}\\b)/);
                    if (m) d.phone = m[1];
                }
                return d;
            }""")

            if extracted.get("rating"): gbp["Rating"] = extracted["rating"]
            if extracted.get("reviews"): gbp["Total Reviews"] = extracted["reviews"]
            if extracted.get("phone"): gbp["Phone"] = extracted["phone"]
            if extracted.get("address"):
                addr = extracted["address"]
                addr = re.sub(r'^[\x00-\x1f\x80-\xff\U000e0000-\U000e007f]+', '', addr).strip()
                gbp["Address"] = addr
            if extracted.get("website"): gbp["Website"] = extracted["website"]
            if extracted.get("hours"): gbp["Business Hours"] = extracted["hours"]

            browser.close()

        def safe_emit(msg):
            try:
                emit(msg)
            except Exception:
                emit(msg.encode('ascii', 'replace').decode('ascii'))

        safe_emit(f"[GBP] Name    : {gbp['Business Name']}")
        safe_emit(f"[GBP] Rating  : {gbp['Rating']}  |  Reviews: {gbp['Total Reviews']}")
        safe_emit(f"[GBP] Phone   : {gbp['Phone']}")
        safe_emit(f"[GBP] Address : {gbp['Address']}")
        safe_emit(f"[GBP] Website : {gbp['Website'][:60] if gbp['Website'] != 'N/A' else 'N/A'}")

    except Exception as e:
        emit(f"[GBP] Error: {str(e)}")

    return gbp


def analyze_website(url: str, business_name: str, location: str,
                    progress_callback=None) -> dict:
    def emit(msg):
        if progress_callback: progress_callback(msg)
        else: print(msg)

    result = {
        "ssl": False, "mobile_friendly": False,
        "meta_title": "", "meta_description": "", "h1_tag": "",
        "sitemap": False, "robots_txt": False, "favicon": False,
        "contact_info_present": False, "google_maps_embedded": False,
        "whatsapp_button": False,
        "location_in_title": False, "location_in_description": False,
        "location_in_h1": False, "location_in_content": False,
        "nap_on_website": False, "contact_page": False,
        "website_url": url or "N/A", "page_reachable": False,
    }

    if not url or url.strip() in ("", "N/A"):
        emit("[WEB] ⚠️ No website URL provided — skipping website analysis.")
        return result

    if not url.startswith("http"):
        url = "https://" + url

    result["ssl"] = url.startswith("https://")
    emit(f"[WEB] Analyzing website: {url}")

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        result["page_reachable"] = (resp.status_code == 200)
        result["ssl"]            = resp.url.startswith("https://")
        result["website_url"]    = resp.url

        if not result["page_reachable"]:
            emit(f"[WEB] ❌ HTTP {resp.status_code} — page not accessible.")
            return result

        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text(separator=" ", strip=True).lower()
        loc  = location.lower()

        result["mobile_friendly"] = bool(soup.find("meta", attrs={"name": "viewport"}))
        emit(f"   SSL: {'✅' if result['ssl'] else '❌'}  |  Mobile: {'✅' if result['mobile_friendly'] else '❌'}")

        t = soup.find("title")
        result["meta_title"] = t.get_text(strip=True) if t else ""

        md = soup.find("meta", attrs={"name": "description"})
        result["meta_description"] = (md.get("content", "") if md else "")

        h1 = soup.find("h1")
        result["h1_tag"] = h1.get_text(strip=True) if h1 else ""
        emit(f"   Title: {'✅' if result['meta_title'] else '❌'}  |  Description: {'✅' if result['meta_description'] else '❌'}  |  H1: {'✅' if result['h1_tag'] else '❌'}")

        result["favicon"] = bool(soup.find("link", rel=lambda r: r and "icon" in " ".join(r).lower()))
        result["google_maps_embedded"] = bool(soup.find("iframe", src=lambda s: s and "google.com/maps" in s))
        result["whatsapp_button"] = bool(soup.find("a", href=lambda h: h and ("wa.me" in h or "api.whatsapp.com" in h)))
        result["contact_page"] = bool(
            soup.find("a", href=lambda h: h and "contact" in h.lower()) or
            soup.find("a", string=re.compile(r"contact", re.I))
        )

        phone_re = re.compile(r'(\+91[\s\-]?\d{10}|\b\d{10}\b|\d{3}[\s\-]\d{3}[\s\-]\d{4}|0\d{2,4}[\s\-]\d{6,8})')
        has_phone = bool(phone_re.search(text))
        result["contact_info_present"] = has_phone

        name_word = (business_name or "").split()[0].lower() if business_name else ""
        has_name  = name_word in text if name_word else False
        has_loc   = loc in text
        result["nap_on_website"] = has_phone and (has_name or has_loc)

        result["location_in_title"]       = loc in result["meta_title"].lower()
        result["location_in_description"] = loc in result["meta_description"].lower()
        result["location_in_h1"]          = loc in result["h1_tag"].lower()
        result["location_in_content"]     = loc in text

        base = f"{urlparse(resp.url).scheme}://{urlparse(resp.url).netloc}"
        try:
            s = requests.get(f"{base}/sitemap.xml", headers=HEADERS, timeout=8)
            result["sitemap"] = (s.status_code == 200 and "xml" in s.text[:200].lower())
        except Exception:
            pass
        try:
            r = requests.get(f"{base}/robots.txt", headers=HEADERS, timeout=8)
            result["robots_txt"] = (r.status_code == 200 and len(r.text.strip()) > 5)
        except Exception:
            pass

    except Exception as e:
        emit(f"[WEB] Error: {e}")

    return result


def calculate_scores(gbp: dict, web: dict) -> dict:
    g = w = l = 0
    if gbp.get("Rating") not in ("N/A", "", None):
        g += 5
        try:
            if float(gbp["Rating"]) >= 4.0: g += 5
        except (ValueError, TypeError): pass
    if gbp.get("Total Reviews") not in ("N/A", "", None):
        try:
            if int(gbp["Total Reviews"]) >= 10: g += 5
        except (ValueError, TypeError): pass
    if gbp.get("Website")  not in ("N/A", "", None): g += 5
    if gbp.get("Phone")    not in ("N/A", "", None): g += 5
    if gbp.get("Address")  not in ("N/A", "", None): g += 5

    if web.get("ssl"):               w += 5
    if web.get("mobile_friendly"):   w += 5
    if web.get("meta_title"):        w += 5
    if web.get("meta_description"):  w += 5
    if web.get("h1_tag"):            w += 5
    if web.get("sitemap"):           w += 5
    if web.get("robots_txt"):        w += 5
    if web.get("favicon"):           w += 5

    if web.get("location_in_title"):       l += 5
    if web.get("location_in_description"): l += 5
    if web.get("location_in_h1"):          l += 5
    if web.get("location_in_content"):     l += 5
    if web.get("google_maps_embedded"):    l += 5
    if web.get("nap_on_website"):          l += 5

    total = g + w + l
    grade = "A" if total >= 85 else "B" if total >= 70 else "C" if total >= 55 else "D" if total >= 40 else "F"

    return {
        "gbp_score": g,   "gbp_max": 30,
        "web_score": w,   "web_max": 40,
        "local_score": l, "local_max": 30,
        "total_score": total, "total_max": 100,
        "grade": grade,
    }


def generate_recommendations(gbp: dict, web: dict, scores: dict) -> list:
    recs = []
    def add(cat, check, status, priority, rec):
        recs.append({"Category": cat, "Check": check, "Status": status,
                     "Priority": priority, "Recommendation": rec})

    if gbp.get("Rating") in ("N/A", "", None):
        add("Google Business Profile", "Rating on GBP", "❌ Missing", "HIGH",
            "Claim your Google Business Profile at business.google.com to display a star rating.")
    else:
        try:
            if float(gbp["Rating"]) < 4.0:
                add("Google Business Profile", "Rating < 4.0 Stars", "⚠️ Low", "HIGH",
                    f"Rating is {gbp['Rating']}★. Actively ask happy customers to leave 5-star reviews via WhatsApp or email.")
        except (ValueError, TypeError): pass

    rev = gbp.get("Total Reviews", "0") or "0"
    try:
        if int(rev) < 10:
            add("Google Business Profile", "Low Review Count", "⚠️ Low", "HIGH",
                "Get more reviews! Share your Google review link with existing customers via WhatsApp, QR code, or SMS.")
    except ValueError: pass

    if gbp.get("Website") in ("N/A", "", None):
        add("Google Business Profile", "Website on GBP", "❌ Missing", "HIGH",
            "Add your website URL to your Google Business Profile to increase credibility.")
    if gbp.get("Phone") in ("N/A", "", None):
        add("Google Business Profile", "Phone on GBP", "❌ Missing", "MEDIUM",
            "Add a phone number so customers can call directly from Google Search.")
    if gbp.get("Address") in ("N/A", "", None):
        add("Google Business Profile", "Address on GBP", "❌ Missing", "HIGH",
            "Add your complete business address so you appear in local map results.")

    if web.get("website_url") in ("N/A", "", None):
        add("Website SEO", "No Website", "❌ Missing", "CRITICAL",
            "Create a website! This is essential for local SEO.")
        return recs

    if not web.get("page_reachable"):
        add("Website SEO", "Website Unreachable", "❌ Down", "CRITICAL",
            "Your website is down. Contact your hosting provider immediately.")
        return recs

    if not web.get("ssl"):
        add("Website SEO", "SSL Certificate (HTTPS)", "❌ Missing", "HIGH",
            "Install SSL. Use Let's Encrypt (free) or enable SSL from your hosting panel.")
    if not web.get("mobile_friendly"):
        add("Website SEO", "Mobile-Friendly", "❌ Missing", "HIGH",
            "Add <meta name='viewport' content='width=device-width, initial-scale=1'> and use responsive CSS.")
    if not web.get("meta_title"):
        add("Website SEO", "Meta Title", "❌ Missing", "HIGH",
            "Add a <title> tag: e.g., 'Best Dentist in Pune | SmileCare Dental Clinic'")
    if not web.get("meta_description"):
        add("Website SEO", "Meta Description", "❌ Missing", "MEDIUM",
            "Add a 150–160 char meta description summarizing your business.")
    if not web.get("h1_tag"):
        add("Website SEO", "H1 Tag", "❌ Missing", "MEDIUM",
            "Add one H1 heading on your homepage.")
    if not web.get("sitemap"):
        add("Website SEO", "Sitemap.xml", "❌ Missing", "MEDIUM",
            "Generate a sitemap at xml-sitemaps.com and upload it.")
    if not web.get("robots_txt"):
        add("Website SEO", "Robots.txt", "❌ Missing", "LOW",
            "Create a robots.txt file at your domain root.")
    if not web.get("favicon"):
        add("Website SEO", "Favicon", "❌ Missing", "LOW",
            "Add a favicon.ico file.")

    if not web.get("location_in_title"):
        add("Local SEO", "Location in Meta Title", "❌ Missing", "HIGH",
            "Include city name in your title tag.")
    if not web.get("location_in_description"):
        add("Local SEO", "Location in Meta Description", "❌ Missing", "HIGH",
            "Mention your city in the meta description.")
    if not web.get("location_in_h1"):
        add("Local SEO", "Location in H1", "❌ Missing", "MEDIUM",
            "Add your city to the H1 tag.")
    if not web.get("location_in_content"):
        add("Local SEO", "Location in Page Content", "❌ Missing", "MEDIUM",
            "Naturally mention your city throughout your page content.")
    if not web.get("google_maps_embedded"):
        add("Local SEO", "Google Maps Embedded", "❌ Missing", "MEDIUM",
            "Embed a Google Maps iframe on your Contact page.")
    if not web.get("nap_on_website"):
        add("Local SEO", "NAP Consistency", "❌ Missing", "HIGH",
            "Display your Name, Address, and Phone (NAP) clearly on your website.")

    if not recs:
        add("Overall", "Excellent SEO!", "✅ All Pass", "NONE",
            "Your local SEO is strong! Continue requesting reviews regularly.")

    return recs


def compare_businesses(client_data: dict, comp_data: dict) -> dict:
    cg = client_data["gbp"]
    cw = client_data["web"]
    cs = client_data["scores"]

    kg = comp_data["gbp"]
    kw = comp_data["web"]
    ks = comp_data["scores"]

    def parse_float(v):
        try: return float(v)
        except Exception: return 0.0

    def parse_int(v):
        try: return int(v)
        except Exception: return 0

    c_rtg, k_rtg = parse_float(cg.get("Rating")), parse_float(kg.get("Rating"))
    c_rev, k_rev = parse_int(cg.get("Total Reviews")), parse_int(kg.get("Total Reviews"))

    def clean_name(name, default):
        if not name or str(name).lower() in ("results", "google maps", "search results", "search", "all", "n/a", ""):
            return default
        return name

    c_name = clean_name(cg.get("Business Name"), client_data.get("input_name", "Client Business"))
    k_name = clean_name(kg.get("Business Name"), comp_data.get("input_name", "Competitor"))

    comparison_rows = [
        {"metric": "Business Name", "client": c_name, "competitor": k_name, "advantage": "—"},
        {"metric": "Overall Score", "client": f"{cs['total_score']} / 100 ({cs['grade']})", "competitor": f"{ks['total_score']} / 100 ({ks['grade']})",
         "advantage": f"🏆 Client (+{cs['total_score']-ks['total_score']} pts)" if cs['total_score'] > ks['total_score'] else (f"🏆 Competitor (+{ks['total_score']-cs['total_score']} pts)" if ks['total_score'] > cs['total_score'] else "🤝 Tied")},
        {"metric": "Google Business Profile Score", "client": f"{cs['gbp_score']} / 30", "competitor": f"{ks['gbp_score']} / 30",
         "advantage": "🏆 Client" if cs['gbp_score'] > ks['gbp_score'] else ("🏆 Competitor" if ks['gbp_score'] > cs['gbp_score'] else "🤝 Tied")},
        {"metric": "Website SEO Score", "client": f"{cs['web_score']} / 40", "competitor": f"{ks['web_score']} / 40",
         "advantage": "🏆 Client" if cs['web_score'] > ks['web_score'] else ("🏆 Competitor" if ks['web_score'] > cs['web_score'] else "🤝 Tied")},
        {"metric": "Local SEO Score", "client": f"{cs['local_score']} / 30", "competitor": f"{ks['local_score']} / 30",
         "advantage": "🏆 Client" if cs['local_score'] > ks['local_score'] else ("🏆 Competitor" if ks['local_score'] > cs['local_score'] else "🤝 Tied")},
        {"metric": "Google Rating", "client": f"{c_rtg}★" if c_rtg else "N/A", "competitor": f"{k_rtg}★" if k_rtg else "N/A",
         "advantage": f"🏆 Client (+{round(c_rtg-k_rtg, 1)}★)" if c_rtg > k_rtg else (f"🏆 Competitor (+{round(k_rtg-c_rtg, 1)}★)" if k_rtg > c_rtg else "🤝 Tied")},
        {"metric": "Total Reviews", "client": str(c_rev), "competitor": str(k_rev),
         "advantage": f"🏆 Client (+{c_rev-k_rev})" if c_rev > k_rev else (f"🏆 Competitor (+{k_rev-c_rev})" if k_rev > c_rev else "🤝 Tied")},
        {"metric": "Phone Number Listed", "client": "Yes" if cg.get("Phone") not in ("N/A","",None) else "No", "competitor": "Yes" if kg.get("Phone") not in ("N/A","",None) else "No", "advantage": "—"},
        {"metric": "Address Listed", "client": "Yes" if cg.get("Address") not in ("N/A","",None) else "No", "competitor": "Yes" if kg.get("Address") not in ("N/A","",None) else "No", "advantage": "—"},
        {"metric": "Website Listed on GBP", "client": "Yes" if cg.get("Website") not in ("N/A","",None) else "No", "competitor": "Yes" if kg.get("Website") not in ("N/A","",None) else "No", "advantage": "—"},
        {"metric": "SSL (HTTPS)", "client": "✅ Enabled" if cw.get("ssl") else "❌ Missing", "competitor": "✅ Enabled" if kw.get("ssl") else "❌ Missing",
         "advantage": "🏆 Client" if cw.get("ssl") and not kw.get("ssl") else ("🏆 Competitor" if kw.get("ssl") and not cw.get("ssl") else "🤝 Tied")},
        {"metric": "Mobile-Friendly", "client": "✅ Pass" if cw.get("mobile_friendly") else "❌ Fail", "competitor": "✅ Pass" if kw.get("mobile_friendly") else "❌ Fail", "advantage": "—"},
        {"metric": "Meta Title", "client": cw.get("meta_title") or "Missing", "competitor": kw.get("meta_title") or "Missing", "advantage": "—"},
        {"metric": "Meta Description", "client": cw.get("meta_description") or "Missing", "competitor": kw.get("meta_description") or "Missing", "advantage": "—"},
        {"metric": "Location in Meta Title", "client": "Yes" if cw.get("location_in_title") else "No", "competitor": "Yes" if kw.get("location_in_title") else "No",
         "advantage": "🏆 Client" if cw.get("location_in_title") and not kw.get("location_in_title") else ("🏆 Competitor" if kw.get("location_in_title") and not cw.get("location_in_title") else "🤝 Tied")},
        {"metric": "Google Maps Embedded", "client": "Yes" if cw.get("google_maps_embedded") else "No", "competitor": "Yes" if kw.get("google_maps_embedded") else "No",
         "advantage": "🏆 Client" if cw.get("google_maps_embedded") and not kw.get("google_maps_embedded") else ("🏆 Competitor" if kw.get("google_maps_embedded") and not cw.get("google_maps_embedded") else "🤝 Tied")},
        {"metric": "NAP on Website", "client": "Yes" if cw.get("nap_on_website") else "No", "competitor": "Yes" if kw.get("nap_on_website") else "No", "advantage": "—"}
    ]

    strengths = []
    weaknesses = []

    if cs['total_score'] > ks['total_score']:
        strengths.append(f"Higher Overall Local SEO Score ({cs['total_score']}/100 vs {ks['total_score']}/100).")
    elif ks['total_score'] > cs['total_score']:
        weaknesses.append(f"Competitor leads in Overall Local SEO Score ({ks['total_score']}/100 vs {cs['total_score']}/100).")

    if c_rtg > k_rtg and c_rtg > 0:
        strengths.append(f"Better Google star rating ({c_rtg}★ vs {k_rtg}★).")
    elif k_rtg > c_rtg and k_rtg > 0:
        weaknesses.append(f"Competitor has a higher Google star rating ({k_rtg}★ vs {c_rtg}★).")

    if c_rev > k_rev:
        strengths.append(f"More total Google reviews ({c_rev} vs {k_rev}).")
    elif k_rev > c_rev:
        weaknesses.append(f"Competitor has {k_rev - c_rev} more Google reviews.")

    if cw.get("ssl") and not kw.get("ssl"):
        strengths.append("Website is secure (HTTPS) whereas competitor website lacks SSL.")
    elif kw.get("ssl") and not cw.get("ssl"):
        weaknesses.append("Competitor has HTTPS security enabled whereas your site lacks SSL.")

    if cw.get("location_in_title") and not kw.get("location_in_title"):
        strengths.append("Meta Title is optimized with target city location keyword.")
    elif kw.get("location_in_title") and not cw.get("location_in_title"):
        weaknesses.append("Competitor targets city location keyword in Meta Title.")

    if cw.get("google_maps_embedded") and not kw.get("google_maps_embedded"):
        strengths.append("Google Maps is embedded on website.")
    elif kw.get("google_maps_embedded") and not cw.get("google_maps_embedded"):
        weaknesses.append("Competitor has Google Maps embedded on their website.")

    if not strengths: strengths.append("Competitive presence across basic listings.")
    if not weaknesses: weaknesses.append("Strong competitive alignment — maintain review velocity.")

    return {
        "rows": comparison_rows,
        "strengths": strengths,
        "weaknesses": weaknesses,
        "client_name": c_name,
        "comp_name": k_name,
        "client_score": cs['total_score'],
        "comp_score": ks['total_score']
    }


def export_to_excel(business_name: str, location: str,
                    gbp: dict, web: dict, scores: dict, recs: list,
                    comp_result: dict = None,
                    progress_callback=None) -> str:
    def emit(msg):
        if progress_callback: progress_callback(msg)
        else: print(msg)

    os.makedirs("output", exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_nm  = re.sub(r"[^\w]", "_", business_name)[:28]
    filepath = os.path.join("output", f"SEO_Report_{safe_nm}_{ts}.xlsx")

    wb = Workbook()

    TEAL    = "006D77"
    TEAL_LT = "E8F8F5"
    NAVY    = "1B4F72"
    WHITE   = "FFFFFF"
    DARK    = "1E2235"
    G_FILL  = "D4EDDA"
    R_FILL  = "F8D7DA"
    Y_FILL  = "FFF3CD"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def hdr(ws, row, col, val, bg=TEAL, fg=WHITE, size=11, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=size, bold=bold, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        return c

    def dat(ws, row, col, val, bg=None, bold=False, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=10, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = thin
        return c

    ws1 = wb.active
    ws1.title = "Business Information"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 55

    ws1.merge_cells("A1:B1")
    c = ws1.cell(row=1, column=1, value="📍  Local SEO Analyzer — Business Information Report")
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.font = Font(name="Calibri", size=15, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:B2")
    c2 = ws1.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y  %I:%M %p')}")
    c2.fill = PatternFill("solid", fgColor=TEAL_LT)
    c2.font = Font(name="Calibri", size=10, color=DARK)
    c2.alignment = Alignment(horizontal="center")

    hdr(ws1, 3, 1, "Field", bg=NAVY)
    hdr(ws1, 3, 2, "Value", bg=NAVY)

    gbp_rows = [
        ("Business Name",    gbp.get("Business Name", "N/A")),
        ("Category",         gbp.get("Category", "N/A")),
        ("Location",         location),
        ("⭐ Rating",         gbp.get("Rating", "N/A")),
        ("💬 Total Reviews",  gbp.get("Total Reviews", "N/A")),
        ("📞 Phone Number",   gbp.get("Phone", "N/A")),
        ("🏠 Address",        gbp.get("Address", "N/A")),
        ("🌐 Website",        gbp.get("Website", "N/A")),
        ("🕐 Business Hours", gbp.get("Business Hours", "Not listed")),
        ("🗺️ Google Maps Link", gbp.get("Google Maps Link", "N/A")),
    ]
    for i, (field, value) in enumerate(gbp_rows, start=4):
        bg = TEAL_LT if i % 2 == 0 else WHITE
        dat(ws1, i, 1, field, bg=bg, bold=True)
        cv = dat(ws1, i, 2, value, bg=bg)
        if "Link" in field and value and value != "N/A":
            cv.hyperlink = value
            cv.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        ws1.row_dimensions[i].height = 20

    sr = len(gbp_rows) + 5
    ws1.merge_cells(f"A{sr}:B{sr}")
    sc = ws1.cell(row=sr, column=1, value="📊  Overall SEO Score Summary")
    sc.fill = PatternFill("solid", fgColor=TEAL)
    sc.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    sc.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[sr].height = 28

    score_grade_color = "D5E8D4" if scores["total_score"] >= 70 else "FFE6CC" if scores["total_score"] >= 50 else R_FILL
    score_rows = [
        ("🏢 Google Business Profile", f"{scores['gbp_score']} / 30"),
        ("🌐 Website SEO Score",       f"{scores['web_score']} / 40"),
        ("📍 Local SEO Score",         f"{scores['local_score']} / 30"),
        ("🏆 OVERALL SCORE",           f"{scores['total_score']} / 100  ·  Grade: {scores['grade']}"),
    ]
    for j, (lbl, val) in enumerate(score_rows, start=sr + 1):
        is_total = j == sr + 4
        bg = score_grade_color if is_total else (TEAL_LT if j % 2 == 0 else WHITE)
        dat(ws1, j, 1, lbl, bg=bg, bold=is_total)
        dat(ws1, j, 2, val, bg=bg, bold=is_total, align="center")
        ws1.row_dimensions[j].height = 22

    ws2 = wb.create_sheet("SEO Audit Results")
    for col, w in zip("ABCD", [30, 45, 16, 12]): ws2.column_dimensions[col].width = w
    ws2.merge_cells("A1:D1")
    t2 = ws2.cell(row=1, column=1, value="🔍  SEO Audit — Detailed Check Results")
    t2.fill = PatternFill("solid", fgColor=TEAL)
    t2.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    for col, label in enumerate(["Category", "Check", "Result", "Points"], start=1):
        hdr(ws2, 2, col, label, bg=NAVY)

    def rv(flag): return "✅ Pass" if flag else "❌ Fail"
    def sp(flag, pts=5): return pts if flag else 0

    try: rtg = float(gbp.get("Rating") or 0)
    except (ValueError, TypeError): rtg = 0
    try: rev_count = int(gbp.get("Total Reviews") or 0)
    except (ValueError, TypeError): rev_count = 0

    audit_rows = [
        ("🏢 Google Business Profile", "Rating on Google Maps", rv(gbp.get("Rating") not in ("N/A","",None)), sp(gbp.get("Rating") not in ("N/A","",None))),
        ("🏢 Google Business Profile", "Rating ≥ 4.0 Stars", "✅ Pass" if rtg >= 4.0 else "⚠️ Low", sp(rtg >= 4.0)),
        ("🏢 Google Business Profile", "Reviews ≥ 10", rv(rev_count >= 10), sp(rev_count >= 10)),
        ("🏢 Google Business Profile", "Website Listed on GBP", rv(gbp.get("Website") not in ("N/A","",None)), sp(gbp.get("Website") not in ("N/A","",None))),
        ("🏢 Google Business Profile", "Phone Number Listed", rv(gbp.get("Phone") not in ("N/A","",None)), sp(gbp.get("Phone") not in ("N/A","",None))),
        ("🏢 Google Business Profile", "Address Listed", rv(gbp.get("Address") not in ("N/A","",None)), sp(gbp.get("Address") not in ("N/A","",None))),
        ("🌐 Website SEO", "SSL (HTTPS)",           rv(web.get("ssl")),             sp(web.get("ssl"))),
        ("🌐 Website SEO", "Mobile-Friendly",       rv(web.get("mobile_friendly")), sp(web.get("mobile_friendly"))),
        ("🌐 Website SEO", "Meta Title",            rv(web.get("meta_title")),      sp(web.get("meta_title"))),
        ("🌐 Website SEO", "Meta Description",      rv(web.get("meta_description")),sp(web.get("meta_description"))),
        ("🌐 Website SEO", "H1 Tag",                rv(web.get("h1_tag")),          sp(web.get("h1_tag"))),
        ("🌐 Website SEO", "Sitemap.xml",           rv(web.get("sitemap")),         sp(web.get("sitemap"))),
        ("🌐 Website SEO", "Robots.txt",            rv(web.get("robots_txt")),      sp(web.get("robots_txt"))),
        ("🌐 Website SEO", "Favicon",               rv(web.get("favicon")),         sp(web.get("favicon"))),
        ("📍 Local SEO", "Location in Meta Title",  rv(web.get("location_in_title")),       sp(web.get("location_in_title"))),
        ("📍 Local SEO", "Location in Meta Desc",   rv(web.get("location_in_description")), sp(web.get("location_in_description"))),
        ("📍 Local SEO", "Location in H1",          rv(web.get("location_in_h1")),          sp(web.get("location_in_h1"))),
        ("📍 Local SEO", "Location in Content",     rv(web.get("location_in_content")),     sp(web.get("location_in_content"))),
        ("📍 Local SEO", "Google Maps Embedded",    rv(web.get("google_maps_embedded")),    sp(web.get("google_maps_embedded"))),
        ("📍 Local SEO", "NAP on Website",          rv(web.get("nap_on_website")),          sp(web.get("nap_on_website"))),
    ]

    for i, (cat, check, result, pts) in enumerate(audit_rows, start=3):
        row_bg = G_FILL if "✅" in result else (Y_FILL if "⚠️" in result else R_FILL)
        dat(ws2, i, 1, cat)
        dat(ws2, i, 2, check)
        dat(ws2, i, 3, result, bg=row_bg, bold=True, align="center")
        dat(ws2, i, 4, pts, bold=True, align="center")
        ws2.row_dimensions[i].height = 20

    tr = len(audit_rows) + 3
    for col in range(1, 5):
        c = ws2.cell(row=tr, column=col)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        c.border = thin
        c.alignment = Alignment(horizontal="center")
    ws2.cell(row=tr, column=1, value="TOTAL SCORE")
    ws2.cell(row=tr, column=2, value=f"Grade: {scores['grade']}")
    ws2.cell(row=tr, column=4, value=f"{scores['total_score']} / 100")

    ws3 = wb.create_sheet("Recommendations")
    for col, w in zip("ABCDE", [25, 32, 16, 12, 65]): ws3.column_dimensions[col].width = w
    ws3.merge_cells("A1:E1")
    t3 = ws3.cell(row=1, column=1, value="💡  SEO Recommendations & Action Plan")
    t3.fill = PatternFill("solid", fgColor=TEAL)
    t3.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    t3.alignment = Alignment(horizontal="center", vertical="center")

    for col, lbl in enumerate(["Category","Issue","Status","Priority","Recommendation"], start=1):
        hdr(ws3, 2, col, lbl, bg=NAVY)

    PRI_CLR = {"CRITICAL":"C0392B","HIGH":"E74C3C","MEDIUM":"F39C12","LOW":"27AE60","NONE":"27AE60"}

    for i, rec in enumerate(recs, start=3):
        pri = rec.get("Priority","LOW")
        dat(ws3, i, 1, rec["Category"])
        dat(ws3, i, 2, rec["Check"])
        dat(ws3, i, 3, rec["Status"])
        pc = ws3.cell(row=i, column=4, value=pri)
        pc.font = Font(name="Calibri", size=10, bold=True, color=PRI_CLR.get(pri,"000000"))
        pc.alignment = Alignment(horizontal="center")
        pc.border = thin
        dat(ws3, i, 5, rec["Recommendation"])
        ws3.row_dimensions[i].height = 32

    if comp_result:
        ws4 = wb.create_sheet("Competitor Comparison")
        for col, w in zip("ABCD", [32, 45, 45, 25]): ws4.column_dimensions[col].width = w
        ws4.merge_cells("A1:D1")
        t4 = ws4.cell(row=1, column=1, value="⚔️  Competitor Local SEO Comparison Report")
        t4.fill = PatternFill("solid", fgColor=TEAL)
        t4.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
        t4.alignment = Alignment(horizontal="center", vertical="center")

        hdr(ws4, 2, 1, "Metric / SEO Factor", bg=NAVY)
        hdr(ws4, 2, 2, f"Client: {comp_result['client_name']}", bg=TEAL)
        hdr(ws4, 2, 3, f"Competitor: {comp_result['comp_name']}", bg="2E4053")
        hdr(ws4, 2, 4, "Advantage / Winner", bg=NAVY)

        for idx, r in enumerate(comp_result["rows"], start=3):
            dat(ws4, idx, 1, r["metric"], bold=True)
            dat(ws4, idx, 2, r["client"])
            dat(ws4, idx, 3, r["competitor"])
            adv_bg = G_FILL if "Client" in r["advantage"] else (R_FILL if "Competitor" in r["advantage"] else WHITE)
            dat(ws4, idx, 4, r["advantage"], bg=adv_bg, bold=True, align="center")

        r_start = len(comp_result["rows"]) + 4
        ws4.merge_cells(f"A{r_start}:D{r_start}")
        st = ws4.cell(row=r_start, column=1, value="🟢  Client Competitive Strengths")
        st.fill = PatternFill("solid", fgColor="27AE60")
        st.font = Font(name="Calibri", size=11, bold=True, color=WHITE)

        cur_row = r_start + 1
        for s_text in comp_result["strengths"]:
            ws4.merge_cells(f"A{cur_row}:D{cur_row}")
            dat(ws4, cur_row, 1, f"•  {s_text}", bg=TEAL_LT)
            cur_row += 1

        cur_row += 1
        ws4.merge_cells(f"A{cur_row}:D{cur_row}")
        wt = ws4.cell(row=cur_row, column=1, value="🔴  Competitor Advantages & Areas to Beat Competitor")
        wt.fill = PatternFill("solid", fgColor="C0392B")
        wt.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cur_row += 1

        for w_text in comp_result["weaknesses"]:
            ws4.merge_cells(f"A{cur_row}:D{cur_row}")
            dat(ws4, cur_row, 1, f"•  {w_text}", bg=R_FILL)
            cur_row += 1

    wb.save(filepath)
    emit(f"[EXCEL] ✅ Saved: {os.path.basename(filepath)}")
    return filepath


def discover_top_competitor(category: str, location: str, client_name: str, progress_callback=None) -> tuple:
    def emit(msg):
        if progress_callback: progress_callback(msg)
        else: print(msg)

    search_query = f"{category} in {location}"
    search_url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
    emit(f"[AUTO-DISCOVERY] 🔍 Searching Google Maps for top local competitor: '{search_query}'...")

    comp_name = ""
    comp_url = ""

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=HEADLESS,
                args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-blink-features=AutomationControlled"]
            )
            ctx = browser.new_context(viewport={"width": 1280, "height": 900}, user_agent=HEADERS["User-Agent"], locale="en-US")
            page = ctx.new_page()
            page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
            time.sleep(3)

            try:
                for txt in ["Accept all", "Reject all", "I agree"]:
                    btn = page.locator(f'button:has-text("{txt}")')
                    if btn.count() > 0:
                        btn.first.click()
                        break
            except Exception:
                pass

            try:
                page.wait_for_selector('a.hfpxzc', timeout=12000)
            except Exception:
                browser.close()
                return "", ""

            cards = page.locator('a.hfpxzc').all()
            client_clean = client_name.lower().strip()

            for card in cards:
                try:
                    aria = (card.get_attribute("aria-label") or "").strip()
                    if not aria:
                        continue

                    # Skip if this card belongs to the client itself
                    if client_clean and (client_clean in aria.lower() or aria.lower() in client_clean):
                        continue

                    comp_name = aria
                    href = card.get_attribute("href") or ""

                    if href:
                        page.goto(href, wait_until="domcontentloaded", timeout=30000)
                        time.sleep(2)
                        website_el = page.locator('[data-item-id="authority"] a, [aria-label^="Website:"] a')
                        if website_el.count() > 0:
                            comp_url = website_el.first.get_attribute("href") or ""
                        if not comp_url:
                            web_text = page.locator('[data-item-id="authority"] .Io6YTe')
                            if web_text.count() > 0:
                                comp_url = web_text.first.inner_text().strip()
                        if comp_url and not comp_url.startswith("http"):
                            comp_url = "https://" + comp_url
                    break
                except Exception:
                    continue

            browser.close()

    except Exception as e:
        emit(f"  [WARN] Auto-competitor discovery warning: {e}")

    if comp_name:
        emit(f"[AUTO-DISCOVERY] 🎯 Discovered top competitor: '{comp_name}' | Website: '{comp_url or 'N/A'}'")
    else:
        emit(f"[AUTO-DISCOVERY] ⚠️ Could not auto-discover competitor for '{category} in {location}'.")

    return comp_name, comp_url


_latest_analysis_data = {}

def get_latest_analysis_data():
    return _latest_analysis_data


def run_analysis(business_name: str, category: str, location: str,
                 website_url: str = "", competitor_name: str = "", competitor_url: str = "",
                 progress_callback=None) -> str:
    global _latest_analysis_data

    def emit(msg):
        if progress_callback: progress_callback(msg)
        else: print(msg)

    emit(f"[START] ▶ Analyzing Client: {business_name} | {category} | {location}")

    gbp    = scrape_gbp(business_name, category, location, progress_callback)
    url    = website_url.strip() or (gbp.get("Website") or "")
    web    = analyze_website(url, business_name, location, progress_callback)
    scores = calculate_scores(gbp, web)
    recs   = generate_recommendations(gbp, web, scores)

    emit(f"[SCORE] 🏆 Client Overall: {scores['total_score']}/100 — Grade: {scores['grade']}")
    emit(f"        GBP: {scores['gbp_score']}/30  |  "
         f"Website: {scores['web_score']}/40  |  "
         f"Local SEO: {scores['local_score']}/30")

    client_data = {"gbp": gbp, "web": web, "scores": scores, "recs": recs, "input_name": business_name}
    comp_result = None

    c_name = competitor_name.strip()
    c_url  = competitor_url.strip()

    if not c_name and not c_url:
        emit(f"\n[AUTO-DISCOVERY] No competitor specified. Searching top local competitor on Google Maps...")
        auto_c_name, auto_c_url = discover_top_competitor(category, location, business_name, progress_callback)
        c_name = auto_c_name
        c_url  = auto_c_url

    if c_name or c_url:
        emit(f"\n[COMPETITOR] ⚔️ Analyzing Competitor: {c_name or c_url}...")
        comp_gbp    = scrape_gbp(c_name or "Competitor", category, location, progress_callback)
        c_final_url = c_url or (comp_gbp.get("Website") or "")
        comp_web    = analyze_website(c_final_url, c_name or comp_gbp.get("Business Name", "Competitor"), location, progress_callback)
        comp_scores = calculate_scores(comp_gbp, comp_web)
        comp_recs   = generate_recommendations(comp_gbp, comp_web, comp_scores)

        comp_data   = {"gbp": comp_gbp, "web": comp_web, "scores": comp_scores, "recs": comp_recs, "input_name": c_name or "Competitor"}
        comp_result = compare_businesses(client_data, comp_data)

        emit(f"[COMPETITOR] ✅ Comparison Complete!")
        emit(f"            Client ({comp_result['client_name']}): {scores['total_score']}/100 vs Competitor ({comp_result['comp_name']}): {comp_scores['total_score']}/100")

    filepath = export_to_excel(business_name, location, gbp, web, scores, recs, comp_result=comp_result, progress_callback=progress_callback)

    _latest_analysis_data = {
        "filepath": filepath,
        "client": client_data,
        "comparison": comp_result
    }

    return filepath
